import openpyxl
from typing import List
from Components.DcdcWidget import DcdcWidget
from Components.PsuWidget import PsuWidget
from Components.LdoWidget import LdoWidget
from Components.SwitchWidget import SwitchWidget
from Components.ConsumerWidget import ConsumerWidget
from Components.Pmic import Pmic
from Formula import Formula

# DCDC DATABASE
DCDC_REF_COMPONENT = 1
DCDC_SUPPLIER = 2
DCDC_CURRENT_MAX = 3
DCDC_MODE = 4
DCDC_EQUIVALENCE_CODE = 5
DCDC_VOLTAGE_INPUT_MIN = 6
DCDC_VOLTAGE_INPUT_MAX = 7
DCDC_VOLTAGE_OUTPUT_MIN = 8
DCDC_VOLTAGE_OUTPUT_MAX = 9
DCDC_EFFICIENCY = 10

# PMIC DATABASE
PMIC_REF_COMPONENT = 1
PMIC_SUPPLIER = 2
PMIC_EQUIVALENCE_CODE = 3
PMIC_COMPONENT = 4
PMIC_CURRENT_MAX = 5
PMIC_VOLTAGE_INPUT_MIN = 6
PMIC_VOLTAGE_INPUT_MAX = 7
PMIC_VOLTAGE_OUTPUT_MIN = 8
PMIC_VOLTAGE_OUTPUT_MAX = 9
PMIC_EFFICIENCY = 10
PMIC_NAME = 11

# DCDC DATABASE
LDO_REF_COMPONENT = 1
LDO_SUPPLIER = 2
LDO_CURRENT_MAX = 3
LDO_EQUIVALENCE_CODE = 4
LDO_VOLTAGE_INPUT_MIN = 5
LDO_VOLTAGE_INPUT_MAX = 6
LDO_VOLTAGE_OUTPUT = 7

# PSU DATABASE
PSU_REF_COMPONENT = 1
PSU_SUPPLIER = 2
PSU_EQUIVALENCE_CODE = 3
PSU_CURRENT_MAX = 4
PSU_VOLTAGE_IN = 5
PSU_VOLTAGE_OUT = 6
PSU_JACK = 7

# SWITCH DATABASE
SWITCH_TYPE = 1
SWITCH_I_LIM = 2
SWITCH_RDS_ON = 3
SWITCH_REF_COMPONENT = 4
SWITCH_SUPPLIER = 5
SWITCH_EQUIVALENCE_CODE = 6
SWITCH_VOLTAGE_IN_MIN = 7
SWITCH_VOLTAGE_IN_MAX = 8
SWITCH_VBIAS_MIN = 9
SWITCH_VBIAS_MAX = 10

# CONSUMER DATABASE
CONSUMER_TYPE = 1
CONSUMER_SUPPLIER = 2
CONSUMER_REF_COMPONENT = 3
CONSUMER_EQUIVALENCE_CODE = 4
CONSUMER_INFO = 5
CONSUMER_VOLTAGE_INPUT = 6
CONSUMER_CURRENT_THEORETICAL = 7
CONSUMER_CURRENT_MIN_MEASURE = 8
CONSUMER_CURRENT_MAX_MEASURE = 9
CONSUMER_CURRENT_PEAK_MEASURE = 10

FILE_DATABASE = "DATA_BASE.xlsx"
FILE_CONSUMER_DATA_BASE = "Consumer_DATA_BASE.xlsx"


def loading_database() -> List[DcdcWidget] and List[PsuWidget] and List[ConsumerWidget]:
    input_file_database = openpyxl.load_workbook(FILE_DATABASE, read_only=True)
    for sheet in input_file_database:
        if sheet.title == "DCDC":
            sheet_dcdc = sheet
        elif sheet.title == "PMIC":
            sheet_pmic = sheet
        elif sheet.title == "PSU":
            sheet_psu = sheet
        elif sheet.title == "LDO":
            sheet_ldo = sheet
        elif sheet.title == "SWITCH":
            sheet_switch = sheet

    print("Loading database ...")

    # Loading DCDC DATABASE
    dcdc_list = []
    line = 1
    for _ in sheet_dcdc:
        line = line + 1
        if str(sheet_dcdc.cell(line, 1).value) != "None":
            ref_component = str(sheet_dcdc.cell(line, DCDC_REF_COMPONENT).value)
            supplier = str(sheet_dcdc.cell(line, DCDC_SUPPLIER).value)
            current_max = float(sheet_dcdc.cell(line, DCDC_CURRENT_MAX).value)
            mode = str(sheet_dcdc.cell(line, DCDC_MODE).value)
            equivalence_code = str(sheet_dcdc.cell(line, DCDC_EQUIVALENCE_CODE).value)
            voltage_input_min = float(sheet_dcdc.cell(line, DCDC_VOLTAGE_INPUT_MIN).value)
            voltage_input_max = float(sheet_dcdc.cell(line, DCDC_VOLTAGE_INPUT_MAX).value)
            voltage_output_min = float(sheet_dcdc.cell(line, DCDC_VOLTAGE_OUTPUT_MIN).value)
            voltage_output_max = float(sheet_dcdc.cell(line, DCDC_VOLTAGE_OUTPUT_MAX).value)
            dcdc_formula = Formula(str(sheet_dcdc.cell(line, DCDC_EFFICIENCY).value))

            dcdc_list.append(DcdcWidget(ref_component, supplier, current_max, mode, equivalence_code, voltage_input_min,
                                        voltage_input_max, voltage_output_min, voltage_output_max, dcdc_formula))

    # Loading PSU DATABASE
    psu_list = []
    line = 1
    for _ in sheet_psu:
        line = line + 1
        if str(sheet_psu.cell(line, 1).value) != "None":
            ref_component = str(sheet_psu.cell(line, PSU_REF_COMPONENT).value)
            supplier = str(sheet_psu.cell(line, PSU_SUPPLIER).value)
            equivalence_code = str(sheet_psu.cell(line, PSU_EQUIVALENCE_CODE).value)
            current_max = float(sheet_psu.cell(line, PSU_CURRENT_MAX).value)
            voltage_input = float(sheet_psu.cell(line, PSU_VOLTAGE_IN).value)
            voltage_output = float(sheet_psu.cell(line, PSU_VOLTAGE_OUT).value)
            jack = str(sheet_psu.cell(line, PSU_JACK).value)

            psu_list.append(PsuWidget(ref_component, supplier, equivalence_code, current_max, voltage_input,
                                      voltage_output, jack))

    # Loading LDO DATABASE
    ldo_list = []
    line = 1
    for _ in sheet_ldo:
        line = line + 1
        if str(sheet_ldo.cell(line, 1).value) != "None":
            ref_component = str(sheet_ldo.cell(line, LDO_REF_COMPONENT).value)
            supplier = str(sheet_ldo.cell(line, LDO_SUPPLIER).value)
            current_max = float(sheet_ldo.cell(line, LDO_CURRENT_MAX).value)
            equivalence_code = str(sheet_ldo.cell(line, LDO_EQUIVALENCE_CODE).value)
            voltage_input_min = float(sheet_ldo.cell(line, LDO_VOLTAGE_INPUT_MIN).value)
            voltage_input_max = float(sheet_ldo.cell(line, LDO_VOLTAGE_INPUT_MAX).value)
            voltage_output = float(sheet_ldo.cell(line, LDO_VOLTAGE_OUTPUT).value)

            ldo_list.append(LdoWidget(ref_component, supplier, current_max, equivalence_code, voltage_input_min,
                                      voltage_input_max, voltage_output))

    # Loading SWITCH DATABASE
    switch_list = []
    line = 1
    for _ in sheet_switch:
        line = line + 1
        if str(sheet_switch.cell(line, 1).value) != "None":
            type = str(sheet_switch.cell(line, SWITCH_TYPE).value)
            current_max = float(sheet_switch.cell(line, SWITCH_I_LIM).value)
            rds_on = float(sheet_switch.cell(line, SWITCH_RDS_ON).value)
            ref_component = str(sheet_switch.cell(line, SWITCH_REF_COMPONENT).value)
            supplier = str(sheet_switch.cell(line, SWITCH_SUPPLIER).value)
            equivalence_code = str(sheet_switch.cell(line, SWITCH_EQUIVALENCE_CODE).value)
            voltage_input_min = float(sheet_switch.cell(line, SWITCH_VOLTAGE_IN_MIN).value)
            voltage_input_max = float(sheet_switch.cell(line, SWITCH_VOLTAGE_IN_MAX).value)
            voltage_bias_min = str(sheet_switch.cell(line, SWITCH_VBIAS_MIN).value)
            voltage_bias_max = str(sheet_switch.cell(line, SWITCH_VBIAS_MAX).value)

            switch_list.append(SwitchWidget(type, current_max, rds_on, ref_component, supplier, equivalence_code,
                                            voltage_input_min, voltage_input_max, voltage_bias_min, voltage_bias_max))

    # Loading PMIC
    pmic_list = []
    dict_pmic = {}
    line = 1
    for _ in sheet_pmic:
        line = line + 1
        if str(sheet_pmic.cell(line, 1).value) != "None":
            ref_component = str(sheet_pmic.cell(line, PMIC_REF_COMPONENT).value)
            supplier = str(sheet_pmic.cell(line, PMIC_SUPPLIER).value)
            equivalence_code = str(sheet_pmic.cell(line, PMIC_EQUIVALENCE_CODE).value)
            component = str(sheet_pmic.cell(line, PMIC_COMPONENT).value)
            current_max = float(sheet_pmic.cell(line, PMIC_CURRENT_MAX).value)
            voltage_input_min = float(sheet_pmic.cell(line, PMIC_VOLTAGE_INPUT_MIN).value)
            voltage_input_max = float(sheet_pmic.cell(line, PMIC_VOLTAGE_INPUT_MAX).value)
            voltage_output_min = float(sheet_pmic.cell(line, PMIC_VOLTAGE_OUTPUT_MIN).value)
            voltage_output_max = float(sheet_pmic.cell(line, PMIC_VOLTAGE_OUTPUT_MAX).value)
            efficiency = float(sheet_pmic.cell(line, PMIC_EFFICIENCY).value)
            name = str(sheet_pmic.cell(line, PMIC_NAME).value)

        if equivalence_code not in dict_pmic:
            new_pmic = Pmic(ref_component, supplier, equivalence_code)
            dict_pmic.update({equivalence_code: new_pmic})
            pmic_list.append(new_pmic)

        if component == "DCDC":
            dcdc_pmic = DcdcWidget(ref_component, supplier, current_max, "Pulse Frequency Mod.", equivalence_code,
                                   voltage_input_min, voltage_input_max, voltage_output_min, voltage_output_max,
                                   Formula("None"))
            dcdc_pmic.efficiency_is_set = True
            dcdc_pmic.efficiency = efficiency
            dcdc_pmic.name = name
            dcdc_pmic.voltage_output = voltage_output_max
            dcdc_pmic.voltage_input = voltage_input_max
            dict_pmic[equivalence_code].add_dcdc(dcdc_pmic)
        elif component == "LDO":
            ldo_pmic = LdoWidget(ref_component, supplier, current_max, equivalence_code, voltage_input_min,
                                 voltage_input_max, voltage_output_max)
            ldo_pmic.name = name
            dict_pmic[equivalence_code].add_ldo(ldo_pmic)
        else:
            print("[DEBUG] Loading PMIC database : unknown component")

    # Loading CONSUMER DATABASE
    input_file_consumer_database = openpyxl.load_workbook(FILE_CONSUMER_DATA_BASE, read_only=True)
    consumer_list = []
    line = 1
    for sheet_consumer in input_file_consumer_database:
        line = 1
        for _ in sheet_consumer:
            line = line + 1
            if str(sheet_consumer.cell(line, 1).value) != "None":
                type = str(sheet_consumer.cell(line, CONSUMER_TYPE).value)
                supplier = str(sheet_consumer.cell(line, CONSUMER_SUPPLIER).value)
                ref_component = str(sheet_consumer.cell(line, CONSUMER_REF_COMPONENT).value)
                equivalence_code = str(sheet_consumer.cell(line, CONSUMER_EQUIVALENCE_CODE).value)
                info = str(sheet_consumer.cell(line, CONSUMER_INFO).value)
                voltage_input = float(sheet_consumer.cell(line, CONSUMER_VOLTAGE_INPUT).value)
                current_theoretical = float(sheet_consumer.cell(line, CONSUMER_CURRENT_THEORETICAL).value)
                current_min_measure = str(sheet_consumer.cell(line, CONSUMER_CURRENT_MIN_MEASURE).value)
                current_max_measure = str(sheet_consumer.cell(line, CONSUMER_CURRENT_MAX_MEASURE).value)
                current_peak_measure = str(sheet_consumer.cell(line, CONSUMER_CURRENT_PEAK_MEASURE).value)

                consumer_list.append(ConsumerWidget(str(sheet_consumer.title), type, supplier, ref_component,
                                                    equivalence_code, info, voltage_input, current_theoretical,
                                                    current_min_measure, current_max_measure, current_peak_measure))

    print("Database loaded !")
    return dcdc_list, psu_list, ldo_list, consumer_list, switch_list, pmic_list
